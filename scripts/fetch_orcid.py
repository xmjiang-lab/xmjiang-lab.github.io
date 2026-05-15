"""
fetch_orcid.py — Jiang Lab @ SISU
=================================
Single command to refresh the website data from ORCID and emit JSON.

Source of truth for publications: ORCID (https://orcid.org/0000-0002-5171-9774).
Anything Prof. Jiang wants on the site — including Chinese papers — should be
added to ORCID first. We do NOT hand-curate publication content here.

What this script does:
  1. Fetches works from every ORCID listed in LAB_ORCIDS.
  2. De-duplicates by DOI (or title prefix if DOI absent).
  3. Opens `site_data.xlsx`. Merges ORCID data into the Publications sheet,
     PRESERVING any manual-only columns (pdf_url, code_url, video_url,
     show, category, group_id, notes). Re-saves the xlsx.
  4. Reads ALL sheets and emits JSON files into assets/data/:
       - publications.json   (sorted by year desc, show != "no")
       - people.json         (current + alumni split, with author_patterns)
       - funders.json        (affiliations / funders for the footer / home strip)
       - projects.json       (research themes, ordered)
       - collaborators.json  (key / emerging collaborators, including bio)
       - news.json           (news items, sorted by date desc)

Run locally (PsychoPy Python works fine):
    pip install requests openpyxl
    python scripts/fetch_orcid.py             # online mode: fetch ORCID + emit
    python scripts/fetch_orcid.py --offline   # xlsx → JSON only, no HTTP

`--offline` mode skips the ORCID fetch / merge / xlsx write-back steps and
emits JSON files straight from the current state of site_data.xlsx. Use it
from the build_content notebook or any time you just edited the xlsx by
hand and want to refresh the JSON without hitting the network.
"""

from __future__ import annotations
import json
import re
import sys
import time
from pathlib import Path
from typing import Any

import requests
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# ── Configure ────────────────────────────────────────────────
# (orcid_id, display name) — add new lab members here.
LAB_ORCIDS = [
    ("0000-0002-5171-9774", "Xiaoming Jiang"),
    # ("0000-0003-3870-5041", "Wenjun Chen"),
]

ROOT = Path(__file__).parent.parent
XLSX = ROOT / "site_data.xlsx"
DATA_DIR = ROOT / "assets" / "data"

# Publications schema
PUB_COLS = [
    "year", "title", "authors", "journal", "type",
    "doi", "url", "pdf_url", "code_url", "video_url",
    "show", "category", "group_id", "source", "notes",
]
PUB_MANUAL_COLS = {"pdf_url", "code_url", "video_url",
                   "show", "category", "group_id", "notes"}
PREPRINT_HOSTS = ["biorxiv.org", "preprints.org", "psyarxiv.com",
                  "arxiv.org", "medrxiv.org", "osf.io"]

HEADER_FILL = PatternFill("solid", fgColor="003F88")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
HEADER_ALIGN = Alignment(horizontal="center", vertical="center")

# Output people.json structure
ROLE_PRIORITY = {
    "pi": 0, "postdoc": 1, "phd": 2, "master": 3,
    "undergrad": 4, "ra": 5, "visiting": 6,
}


# ── ORCID fetch ──────────────────────────────────────────────
def fetch_orcid_works(orcid_id: str) -> list[dict[str, Any]]:
    """Pull all works for one ORCID via the public v3 API."""
    url = f"https://pub.orcid.org/v3.0/{orcid_id}/works"
    print(f"  fetching {url}")
    r = requests.get(url, headers={"Accept": "application/json"}, timeout=30)
    r.raise_for_status()
    summary = r.json()

    out = []
    for group in summary.get("group", []):
        works = group.get("work-summary") or []
        if not works:
            continue
        s = works[0]
        put_code = s.get("put-code")

        # Fetch the full work record to get authors
        try:
            detail_url = f"https://pub.orcid.org/v3.0/{orcid_id}/work/{put_code}"
            detail = requests.get(
                detail_url, headers={"Accept": "application/json"}, timeout=30
            ).json()
        except Exception:
            detail = s

        title = ((s.get("title") or {}).get("title") or {}).get("value") or ""
        year = ""
        pub_date = s.get("publication-date") or {}
        if pub_date.get("year"):
            year = pub_date["year"].get("value") or ""

        doi = ""
        url_val = ""
        for eid in ((s.get("external-ids") or {}).get("external-id") or []):
            if (eid.get("external-id-type") or "").lower() == "doi":
                doi = (eid.get("external-id-value") or "").strip()
                url_val = (eid.get("external-id-url") or {}).get("value") or f"https://doi.org/{doi}"
                break
        if not url_val:
            # fallback: any external id that has a URL
            for eid in ((s.get("external-ids") or {}).get("external-id") or []):
                u = (eid.get("external-id-url") or {}).get("value")
                if u:
                    url_val = u
                    break

        # Authors
        names = []
        for c in ((detail.get("contributors") or {}).get("contributor") or []):
            cn = (c.get("credit-name") or {}).get("value")
            if cn:
                names.append(cn)
        authors = "; ".join(names)

        journal = ((s.get("journal-title") or {}).get("value")) or ""
        typ = (s.get("type") or "").lower().replace("_", "-")

        out.append({
            "year": year,
            "title": title.strip(),
            "authors": authors,
            "journal": journal,
            "type": typ,
            "doi": doi,
            "url": url_val,
            "pdf_url": "",
            "code_url": "",
            "video_url": "",
            "show": "yes",
            "category": "",
            "group_id": "",
            "source": orcid_id,
            "notes": "",
        })
        time.sleep(0.05)  # be polite

    print(f"    -> {len(out)} works")
    return out


def pub_key(rec: dict[str, Any]) -> tuple[str, str]:
    """De-dup key. Prefer DOI; fall back to title prefix."""
    doi = (rec.get("doi") or "").strip().lower()
    if doi:
        return ("doi", doi)
    title = (rec.get("title") or "").strip().lower()
    return ("title", re.sub(r"\s+", " ", title)[:60])


# ── xlsx I/O ─────────────────────────────────────────────────
def read_pubs_sheet(xlsx_path: Path) -> dict[tuple, dict]:
    if not xlsx_path.exists():
        return {}
    wb = load_workbook(xlsx_path)
    if "Publications" not in wb.sheetnames:
        return {}
    ws = wb["Publications"]
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 2:
        return {}
    headers = [str(h) if h is not None else "" for h in rows[0]]
    out = {}
    # Row 2 is the yellow "comment" instruction row — skip it.
    # We detect it by checking the first cell for a known instruction phrase
    # or by checking if any cell starts with "auto" / "MANUAL".
    start = 1
    if rows[1] and any(
        isinstance(c, str) and (c.startswith("auto") or c.startswith("MANUAL") or "yes/no" in c)
        for c in rows[1] if c
    ):
        start = 2
    for row in rows[start:]:
        if not any(row):
            continue
        rec = {h: ("" if v is None else v) for h, v in zip(headers, row)}
        out[pub_key(rec)] = rec
    return out


def write_pubs_sheet(xlsx_path: Path, records: list[dict]):
    """Re-emit the Publications sheet with header + comment row + records."""
    if xlsx_path.exists():
        wb = load_workbook(xlsx_path)
    else:
        wb = Workbook()
        wb.remove(wb.active)

    if "Publications" in wb.sheetnames:
        del wb["Publications"]
    # Insert as first sheet
    ws = wb.create_sheet("Publications", 0)

    # Header row
    for i, h in enumerate(PUB_COLS, start=1):
        c = ws.cell(row=1, column=i, value=h)
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        c.alignment = HEADER_ALIGN

    # Comment row
    yellow = PatternFill("solid", fgColor="FCB813")
    italic = Font(color="1A1A1A", bold=True, italic=True, size=10)
    comments = {
        "year": "auto",
        "title": "auto (ORCID)",
        "authors": "auto (ORCID)",
        "journal": "auto (ORCID)",
        "type": "auto (ORCID)",
        "doi": "auto (ORCID)",
        "url": "auto (ORCID)",
        "pdf_url": "MANUAL",
        "code_url": "MANUAL",
        "video_url": "MANUAL",
        "show": "yes/no",
        "category": "MANUAL override",
        "group_id": "MANUAL",
        "source": "auto (orcid id)",
        "notes": "MANUAL: badge text",
    }
    for i, h in enumerate(PUB_COLS, start=1):
        c = ws.cell(row=2, column=i, value=comments.get(h, ""))
        c.fill = yellow
        c.font = italic

    # Data rows
    for r_idx, rec in enumerate(records, start=3):
        for c_idx, col in enumerate(PUB_COLS, start=1):
            ws.cell(row=r_idx, column=c_idx, value=rec.get(col, ""))

    # Column widths
    widths = [6, 60, 40, 30, 16, 25, 40, 35, 35, 35, 6, 12, 12, 8, 22]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A3"
    ws.row_dimensions[1].height = 22
    ws.row_dimensions[2].height = 24

    wb.save(xlsx_path)


def read_sheet_records(xlsx_path: Path, sheet_name: str) -> list[dict]:
    """Read a sheet, skip header row + comment row, return list of dicts."""
    if not xlsx_path.exists():
        return []
    wb = load_workbook(xlsx_path, read_only=True)
    if sheet_name not in wb.sheetnames:
        return []
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 2:
        return []
    headers = [str(h) if h is not None else "" for h in rows[0]]
    # Skip the comment/instruction row if present
    start = 1
    if rows[1] and any(
        isinstance(c, str) and ("MANUAL" in c or "auto" in c or "yes/no" in c or
                                "Family, Given" in c or "stable slug" in c or
                                "filename" in c or "Chinese" in c or "English" in c or
                                "display order" in c or "Hover tooltip" in c)
        for c in rows[1] if c
    ):
        start = 2
    out = []
    for row in rows[start:]:
        if not any(row):
            continue
        rec = {h: ("" if v is None else v) for h, v in zip(headers, row)}
        if str(rec.get("show", "yes")).lower() in ("no", "false", "0", ""):
            if rec.get("show") not in (None, ""):
                continue
        out.append(rec)
    return out


# ── Merging logic ────────────────────────────────────────────
def merge_pubs(existing: dict, fetched: list[dict]) -> list[dict]:
    """
    Refresh ORCID-sourced columns; keep manual columns sacred.

    Strategy:
    - For each fetched record, if key exists, copy NON-manual fields onto the
      existing record. Manual fields remain whatever they were in xlsx.
    - If key doesn't exist, add as new with show=yes.
    - We DON'T delete records that exist in xlsx but not in ORCID — that's
      the path for adding entries that ORCID doesn't have (e.g. Chinese
      journals before Prof. Jiang adds them to ORCID).
    """
    merged = dict(existing)
    title_index = {
        re.sub(r"\s+", " ", str(r.get("title", "")).strip().lower())[:60]: k
        for k, r in merged.items()
    }

    for rec in fetched:
        key = pub_key(rec)
        if key in merged:
            cur = merged[key]
            for col in PUB_COLS:
                if col in PUB_MANUAL_COLS:
                    continue
                if rec.get(col) not in ("", None):
                    cur[col] = rec[col]
            merged[key] = cur
        else:
            # Try title-prefix match (handles DOI-vs-no-DOI for same paper)
            tp = re.sub(r"\s+", " ", str(rec.get("title", "")).strip().lower())[:60]
            if tp and tp in title_index:
                k = title_index[tp]
                cur = merged[k]
                for col in PUB_COLS:
                    if col in PUB_MANUAL_COLS:
                        continue
                    if rec.get(col) not in ("", None) and not cur.get(col):
                        cur[col] = rec[col]
                continue
            merged[key] = rec
            if tp:
                title_index[tp] = key

    return list(merged.values())


def resolve_category(p: dict) -> str:
    """Compute the category badge: journal / preprint / conference / chinese / other."""
    if p.get("category"):
        return str(p["category"]).lower().strip()
    u = (p.get("url") or "").lower()
    if any(h in u for h in PREPRINT_HOSTS):
        return "preprint"
    t = (p.get("type") or "").lower().replace("_", "-")
    if t == "journal-article":  return "journal"
    if t == "conference-paper": return "conference"
    if t == "preprint":         return "preprint"
    return "other"


# ── JSON emitters ─────────────────────────────────────────────
def emit_publications(records: list[dict]):
    """Write assets/data/publications.json — visible, sorted by year desc."""
    visible = [
        r for r in records
        if str(r.get("show", "yes")).lower() not in ("no", "false", "0")
    ]
    # Resolve missing categories
    for r in visible:
        if not r.get("category"):
            r["category"] = resolve_category(r)
    # Sort: year desc, with "" (in press) at top
    visible.sort(key=lambda r: (str(r.get("year", "") or "0"),), reverse=True)
    # Clean empty fields for compact JSON
    out = [
        {k: v for k, v in r.items() if v not in (None, "", "None")}
        for r in visible
    ]
    target = DATA_DIR / "publications.json"
    target.write_text(
        json.dumps(out, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    print(f"  -> {target.name} ({len(out)} entries)")


def emit_people(rows: list[dict]):
    """Write assets/data/people.json — current + alumni split, with author_patterns."""
    # Build author_patterns from each member's English name
    patterns = []
    for r in rows:
        name_en = (r.get("name_en") or "").strip()
        if not name_en:
            continue
        parts = name_en.split()
        if len(parts) >= 2:
            given, family = parts[0], parts[-1]
            given_initial = given[0]
            patterns.extend([
                f"{re.escape(family)},\\s?{re.escape(given_initial)}\\.?(\\s?[A-Z]\\.?)?",
                f"{re.escape(family)},\\s?{re.escape(given)}",
                f"{re.escape(given_initial)}\\.?\\s?[A-Z]?\\.?\\s?{re.escape(family)}",
                f"{re.escape(given)}\\s{re.escape(family)}",
            ])

    def by_role(r):
        return (
            ROLE_PRIORITY.get((r.get("role") or "").lower(), 99),
            -int(str(r.get("cohort_year") or "0") or "0"),
        )

    current = sorted(
        [r for r in rows if (r.get("status") or "").lower() == "current"],
        key=by_role,
    )
    alumni = sorted(
        [r for r in rows if (r.get("status") or "").lower() == "alumni"],
        key=lambda r: -int(str(r.get("cohort_year") or "0") or "0"),
    )

    def serialize(r):
        out = {
            "id": r.get("id") or "",
            "role": r.get("role") or "",
            "cohort_year": str(r.get("cohort_year") or ""),
            "name_zh": r.get("name_zh") or "",
            "name_en": r.get("name_en") or "",
            "title_zh": r.get("title_zh") or "",
            "title_en": r.get("title_en") or "",
            "secondary_title_zh": r.get("secondary_title_zh") or "",
            "secondary_title_en": r.get("secondary_title_en") or "",
            "affil_zh": r.get("affil_zh") or "",
            "affil_en": r.get("affil_en") or "",
            "bio_zh": r.get("bio_zh") or "",
            "bio_en": r.get("bio_en") or "",
            "education": r.get("education") or "",
            "education_zh": r.get("education_zh") or "",
            "research_areas": r.get("research_areas") or "",
            "research_areas_zh": r.get("research_areas_zh") or "",
            "photo": r.get("photo") or "",
            "email": r.get("email") or "",
            "orcid": r.get("orcid") or "",
            "scholar": r.get("scholar") or "",
            "researchgate": r.get("researchgate") or "",
            "homepage": r.get("homepage") or "",
            "now": r.get("now") or "",
            "period": r.get("period") or "",
            "end_date": r.get("end_date") or "",
            "next_position": r.get("next_position") or "",
        }
        return {k: v for k, v in out.items() if v not in ("", None)}

    payload = {
        "author_patterns": patterns,
        "current": [serialize(r) for r in current],
        "alumni":  [serialize(r) for r in alumni],
    }
    target = DATA_DIR / "people.json"
    target.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"  -> {target.name} ({len(current)} current, {len(alumni)} alumni)")


def _order_key(r: dict) -> int:
    """Robust integer order key. Treat blank/None as last; 0 stays 0."""
    v = r.get("order")
    if v is None or v == "":
        return 999
    try:
        return int(v)
    except (TypeError, ValueError):
        return 999


def emit_news(rows: list[dict]):
    """Write news.json — sorted by date desc, featured flag preserved."""
    def _date_key(r):
        return str(r.get("date") or "")

    rows_sorted = sorted(rows, key=_date_key, reverse=True)
    items = []
    for r in rows_sorted:
        featured = str(r.get("featured") or "").lower() in ("yes", "true", "1")
        item = {
            "date": r.get("date") or "",
            "featured": True if featured else None,
            "tag_zh": r.get("tag_zh") or "",
            "tag_en": r.get("tag_en") or "",
            "title_zh": r.get("title_zh") or "",
            "title_en": r.get("title_en") or "",
            "summary_zh": r.get("summary_zh") or "",
            "summary_en": r.get("summary_en") or "",
            "body_zh": r.get("body_zh") or "",
            "body_en": r.get("body_en") or "",
            "link": r.get("link") or "",
        }
        items.append({k: v for k, v in item.items() if v not in ("", None)})
    payload = {"items": items}
    target = DATA_DIR / "news.json"
    target.write_text(
        json.dumps(payload, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    print(f"  -> {target.name} ({len(items)} entries)")


def emit_collaborators(rows: list[dict]):
    """Write collaborators.json — preserves order, includes bio when populated.
    The `type` column distinguishes senior vs early_career collaborators."""
    rows_sorted = sorted(rows, key=_order_key)
    out = []
    for r in rows_sorted:
        out.append({k: v for k, v in {
            "type": r.get("type") or "",
            "name": r.get("name") or "",
            "affiliation": r.get("affiliation") or "",
            "url": r.get("url") or "",
            "bio": r.get("bio") or "",
        }.items() if v})
    target = DATA_DIR / "collaborators.json"
    target.write_text(json.dumps(out, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"  -> {target.name} ({len(out)} entries)")


def emit_funders(rows: list[dict]):
    out = []
    for r in rows:
        out.append({k: v for k, v in {
            "name": r.get("name") or "",
            "logo_path": r.get("logo_path") or "",
            "tooltip": r.get("tooltip") or "",
            "period": r.get("period") or "",
            "url": r.get("url") or "",
        }.items() if v})
    target = DATA_DIR / "funders.json"
    target.write_text(json.dumps(out, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"  -> {target.name} ({len(out)} entries)")


def emit_projects(rows: list[dict]):
    rows_sorted = sorted(rows, key=_order_key)
    out = []
    for r in rows_sorted:
        out.append({k: v for k, v in {
            "id": r.get("id") or "",
            "icon": r.get("icon") or "",
            "title_zh": r.get("title_zh") or "",
            "title_en": r.get("title_en") or "",
            "body_zh": r.get("body_zh") or "",
            "body_en": r.get("body_en") or "",
        }.items() if v})
    target = DATA_DIR / "projects.json"
    target.write_text(json.dumps(out, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"  -> {target.name} ({len(out)} entries)")


# ── Main ────────────────────────────────────────────────────
def main():
    offline = "--offline" in sys.argv[1:]
    mode_tag = " (offline mode — xlsx only)" if offline else ""
    print(f"Jiang Lab @ SISU — data sync{mode_tag}")
    print("=" * 60)
    DATA_DIR.mkdir(parents=True, exist_ok=True)

    if not XLSX.exists():
        print(f"ERROR: {XLSX} not found. Run `python scripts/build_xlsx.py` first.")
        sys.exit(1)

    # 1) Read existing Publications sheet (preserves manual fields)
    existing = read_pubs_sheet(XLSX)
    print(f"Existing publications in xlsx: {len(existing)}")

    if offline:
        # read_pubs_sheet returns a dict[key, record]; emit_publications expects
        # a list of record dicts (same shape merge_pubs returns).
        pubs = list(existing.values())
        print(f"\nSkipping ORCID fetch. Emitting JSON straight from xlsx.")
    else:
        # 2) Fetch ORCID
        fetched = []
        print(f"\nFetching ORCID:")
        for orcid_id, name in LAB_ORCIDS:
            print(f"  {name} ({orcid_id})")
            try:
                fetched.extend(fetch_orcid_works(orcid_id))
            except Exception as e:
                print(f"    !! skipped: {e}")

        # 3) Merge
        merged = merge_pubs(existing, fetched)
        print(f"\nMerged total: {len(merged)} publications")

        # 4) Write back to xlsx
        write_pubs_sheet(XLSX, merged)
        print(f"  -> {XLSX.name} updated")
        pubs = merged

    # 5) Emit all JSON
    print(f"\nEmitting JSON to {DATA_DIR}/:")
    emit_publications(pubs)
    emit_people(read_sheet_records(XLSX, "People"))
    emit_funders(read_sheet_records(XLSX, "Funders"))
    emit_projects(read_sheet_records(XLSX, "Projects"))
    emit_collaborators(read_sheet_records(XLSX, "Collaborators"))
    emit_news(read_sheet_records(XLSX, "News"))

    print(f"\nDone.")


if __name__ == "__main__":
    main()
