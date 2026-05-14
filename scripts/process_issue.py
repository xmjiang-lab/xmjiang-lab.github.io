"""
process_issue.py — Jiang Lab @ SISU
====================================
Called from .github/workflows/process-approved-issue.yml after Prof. Jiang
adds the `approved` label to a form-submission Issue.

Steps:
  1. Read the Issue body (via `gh` CLI; the workflow exports the body to a
     temp file, or we fetch it with `gh issue view`).
  2. Parse the YAML-like block between `---` markers.
  3. Dispatch on `form_type`:
       student-application -> add a row to People sheet, save photo
       graduation          -> flip status to alumni, fill end_date + next_position
       news                -> append a row to News sheet
  4. Re-emit the affected JSON files (no ORCID fetch).

Usage:
    python scripts/process_issue.py <issue_number>
"""
from __future__ import annotations

import base64
import json
import re
import subprocess
import sys
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).parent.parent
XLSX = ROOT / "site_data.xlsx"
DATA_DIR = ROOT / "assets" / "data"
PEOPLE_PHOTO_DIR = ROOT / "assets" / "images" / "people"

sys.path.insert(0, str(Path(__file__).parent))
import fetch_orcid as fo  # noqa: E402


# ── Issue body parsing ──────────────────────────────────────────────────────

YAML_FENCE = re.compile(r"^---\s*$", re.MULTILINE)
PHOTO_FENCE = re.compile(r"```photo_base64\n(.*?)\n```", re.DOTALL)
KV_LINE = re.compile(r'^([a-z_][a-z0-9_]*)\s*:\s*"((?:[^"\\]|\\.)*)"\s*$')


def parse_issue_body(body: str) -> tuple[dict, str | None]:
    """Return (fields_dict, photo_base64_or_None)."""
    fences = list(YAML_FENCE.finditer(body))
    if len(fences) < 2:
        raise ValueError("Issue body missing `---` YAML fences")
    start, end = fences[0].end(), fences[1].start()
    block = body[start:end]

    fields: dict[str, str] = {}
    for line in block.splitlines():
        line = line.strip()
        if not line:
            continue
        m = KV_LINE.match(line)
        if m:
            key, val = m.group(1), m.group(2)
            val = val.replace("\\n", "\n").replace('\\"', '"')
            fields[key] = val
            continue
        # Support unquoted simple `key: value` (e.g. form_type: student-application)
        if ":" in line:
            k, _, v = line.partition(":")
            fields[k.strip()] = v.strip()

    photo_b64 = None
    pm = PHOTO_FENCE.search(body)
    if pm:
        photo_b64 = re.sub(r"\s+", "", pm.group(1))

    return fields, photo_b64


# ── xlsx helpers ────────────────────────────────────────────────────────────

def find_header_col(ws, name: str) -> int:
    for i, c in enumerate(ws[1], start=1):
        if (c.value or "") == name:
            return i
    raise KeyError(f"Column not found: {name}")


def append_row(ws, row_dict: dict[str, str]):
    """Append a new row after the last non-empty row, mapping by header name."""
    headers = [(c.value or "") for c in ws[1]]
    new_row = [row_dict.get(h, "") for h in headers]
    ws.append(new_row)


# ── Slug generation ────────────────────────────────────────────────────────

_SLUG_RE = re.compile(r"[^a-z0-9]+")


def slugify(s: str) -> str:
    s = s.lower().strip()
    s = _SLUG_RE.sub("-", s).strip("-")
    return s


def make_member_id(name_en: str, name_zh: str, issue_number: int) -> str:
    if name_en:
        slug = slugify(name_en)
        if slug:
            return slug
    return f"applicant-issue-{issue_number}"


# ── Handlers ────────────────────────────────────────────────────────────────

def handle_student_application(fields: dict, photo_b64: str | None, issue_n: int):
    wb = load_workbook(XLSX)
    ws = wb["People"]

    member_id = make_member_id(
        fields.get("name_en", ""),
        fields.get("name_zh", ""),
        issue_n,
    )

    photo_filename = ""
    if photo_b64:
        try:
            data = base64.b64decode(photo_b64)
            PEOPLE_PHOTO_DIR.mkdir(parents=True, exist_ok=True)
            photo_filename = f"{member_id}.jpg"
            (PEOPLE_PHOTO_DIR / photo_filename).write_bytes(data)
        except Exception as e:
            print(f"WARNING: could not decode photo: {e}")
            photo_filename = ""

    row = {
        "id": member_id,
        "status": "current",
        "role": fields.get("layer", ""),
        "cohort_year": fields.get("cohort_year", ""),
        "name_zh": fields.get("name_zh", ""),
        "name_en": fields.get("name_en", ""),
        "title_zh": "",
        "title_en": "",
        "affil_zh": "",
        "affil_en": "",
        "bio_zh": fields.get("bio_zh", ""),
        "bio_en": fields.get("bio_en", ""),
        "education": fields.get("education", ""),
        "research_areas": fields.get("research_areas", ""),
        "photo": photo_filename,
        "email": fields.get("email", ""),
        "orcid": fields.get("orcid", ""),
        "scholar": fields.get("scholar", ""),
        "homepage": fields.get("homepage", ""),
        "now": "",
        "period": "",
        "end_date": "",
        "next_position": "",
        "show": "yes",
    }
    append_row(ws, row)
    wb.save(XLSX)
    print(f"  Added People row: {member_id} ({fields.get('name_zh','')})")


def handle_graduation(fields: dict):
    name_zh = (fields.get("name_zh") or "").strip()
    if not name_zh:
        raise ValueError("graduation: name_zh is required")

    wb = load_workbook(XLSX)
    ws = wb["People"]

    name_col   = find_header_col(ws, "name_zh")
    status_col = find_header_col(ws, "status")
    end_col    = find_header_col(ws, "end_date")
    next_col   = find_header_col(ws, "next_position")
    period_col = find_header_col(ws, "period")
    now_col    = find_header_col(ws, "now")

    target_row = None
    for r in range(3, ws.max_row + 1):
        v = ws.cell(row=r, column=name_col).value
        if v and v.strip() == name_zh:
            target_row = r
            break
    if not target_row:
        raise ValueError(f"graduation: no current member found with name_zh={name_zh!r}")

    next_zh = fields.get("next_position_zh", "")
    next_en = fields.get("next_position_en", "")
    next_combined = next_zh
    if next_en:
        next_combined = f"{next_zh} / {next_en}" if next_zh else next_en

    ws.cell(row=target_row, column=status_col, value="alumni")
    ws.cell(row=target_row, column=end_col,    value=fields.get("end_date", ""))
    ws.cell(row=target_row, column=next_col,   value=next_combined)
    if fields.get("period_summary"):
        ws.cell(row=target_row, column=period_col, value=fields["period_summary"])
    # Also surface `now` for the alumni card renderer.
    if next_combined and not ws.cell(row=target_row, column=now_col).value:
        ws.cell(row=target_row, column=now_col, value=next_combined)

    wb.save(XLSX)
    print(f"  Moved {name_zh} to alumni (row {target_row})")


def handle_news(fields: dict):
    wb = load_workbook(XLSX)
    ws = wb["News"]
    row = {
        "date": fields.get("date", ""),
        "featured": "",
        "tag_zh": fields.get("tag_zh", ""),
        "tag_en": fields.get("tag_en", ""),
        "title_zh": fields.get("title_zh", ""),
        "title_en": fields.get("title_en", ""),
        "summary_zh": fields.get("summary_zh", "") or fields.get("body_zh", "")[:80],
        "summary_en": fields.get("summary_en", "") or fields.get("body_en", "")[:120],
        "body_zh": fields.get("body_zh", ""),
        "body_en": fields.get("body_en", ""),
        "link": fields.get("link", ""),
        "show": "yes",
    }
    append_row(ws, row)
    wb.save(XLSX)
    print(f"  Added News row: {row['title_zh'] or row['title_en']}")


# ── Re-emit JSON ────────────────────────────────────────────────────────────

def regenerate_json(form_type: str):
    """Re-emit only the JSON files affected by this form_type."""
    if form_type in ("student-application", "graduation"):
        fo.emit_people(fo.read_sheet_records(XLSX, "People"))
    if form_type == "news":
        fo.emit_news(fo.read_sheet_records(XLSX, "News"))


# ── Main ────────────────────────────────────────────────────────────────────

def fetch_issue_body(issue_number: int) -> str:
    """Use `gh issue view` to fetch the issue body."""
    res = subprocess.run(
        ["gh", "issue", "view", str(issue_number), "--json", "body"],
        capture_output=True,
        text=True,
        check=True,
    )
    return json.loads(res.stdout)["body"] or ""


def main():
    if len(sys.argv) != 2:
        print("Usage: python scripts/process_issue.py <issue_number>", file=sys.stderr)
        sys.exit(2)
    issue_number = int(sys.argv[1])

    print(f"Fetching issue #{issue_number}")
    body = fetch_issue_body(issue_number)

    fields, photo_b64 = parse_issue_body(body)
    form_type = fields.get("form_type", "").strip()
    if not form_type:
        raise ValueError("Issue body missing `form_type`")

    print(f"form_type: {form_type}")
    if form_type == "student-application":
        handle_student_application(fields, photo_b64, issue_number)
    elif form_type == "graduation":
        handle_graduation(fields)
    elif form_type == "news":
        handle_news(fields)
    else:
        raise ValueError(f"Unknown form_type: {form_type}")

    regenerate_json(form_type)
    print("Done.")


if __name__ == "__main__":
    try:
        main()
    except Exception as e:
        print(f"ERROR: {e}", file=sys.stderr)
        sys.exit(1)
